import os
import openpyxl
from openpyxl.utils import range_boundaries

# Configuration
SOURCE_DIR = r"d:\GoogleAntigravityProjects\datacenter document\原始资料\01 需求匹配"
OUTPUT_DIR = r"d:\GoogleAntigravityProjects\datacenter document\output_LLM"
VERSION_SUFFIX = "_v1.0.md"

def ensure_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)

def get_merged_cell_map(sheet):
    """
    Returns a map of (row, col) -> (rowspan, colspan, is_hidden)
    """
    merged_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        
        # The top-left cell gets the rowspan/colspan
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        merged_map[(min_row, min_col)] = {'rowspan': rowspan, 'colspan': colspan, 'hidden': False}
        
        # Mark other cells in the range as hidden
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                merged_map[(row, col)] = {'hidden': True}
                
    return merged_map

def escape_html(text):
    if text is None:
        return ""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def  sheet_to_html_table(sheet):
    merged_map = get_merged_cell_map(sheet)
    html = ["<table>"]
    
    # Iterate through rows
    # openpyxl rows are 1-based
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        # We assume the first row is header if we want specific styling, 
        # but for generic conversion we might just use <tr><td>
        # However, typically first row is header. Let's try to detect or just treat all as td/th if row_idx==1
        
        tr_tag = "  <tr>"
        row_content = []
        is_row_empty = True
        
        for col_idx, cell in enumerate(row, start=1):
            cell_info = merged_map.get((row_idx, col_idx))
            
            if cell_info and cell_info.get('hidden'):
                continue
                
            value = cell.value
            if value is not None:
                is_row_empty = False
            
            tag = "td"
            if row_idx == 1: # Simple heuristic for header
                tag = "th"
                
            attrs = ""
            if cell_info:
                if cell_info['rowspan'] > 1:
                    attrs += f' rowspan="{cell_info["rowspan"]}"'
                if cell_info['colspan'] > 1:
                    attrs += f' colspan="{cell_info["colspan"]}"'
            
            escaped_value = escape_html(value).replace("\n", "<br>")
            row_content.append(f'    <{tag}{attrs}>{escaped_value}</{tag}>')
            
        if not is_row_empty:
            html.append(tr_tag)
            html.extend(row_content)
            html.append("  </tr>")
            
    html.append("</table>")
    return "\n".join(html)

def escape_md(text):
    if text is None:
        return ""
    # Basic escaping for markdown table cells
    return str(text).replace("|", "\\|").replace("\n", "<br>")

def sheet_to_md_table(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""
        
    md_lines = []
    
    # Header
    header = rows[0]
    md_Header = "| " + " | ".join(escape_md(h) for h in header) + " |"
    md_lines.append(md_Header)
    
    # Separator
    md_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
    
    # Body
    for row in rows[1:]:
        md_row = "| " + " | ".join(escape_md(c) for c in row) + " |"
        md_lines.append(md_row)
        
    return "\n".join(md_lines)

def convert_file(filepath):
    filename = os.path.basename(filepath)
    name_without_ext = os.path.splitext(filename)[0]
    output_filename = f"{name_without_ext}{VERSION_SUFFIX}"
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    
    print(f"Converting {filename}...")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        md_content = []
        
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            md_content.append(f"# {sheetname}\n")
            
            # Check for merged cells
            if sheet.merged_cells:
                print(f"  Sheet '{sheetname}' has merged cells. Using HTML.")
                table_content = sheet_to_html_table(sheet)
            else:
                print(f"  Sheet '{sheetname}' is standard. Using Markdown.")
                table_content = sheet_to_md_table(sheet)
                
            md_content.append(table_content)
            md_content.append("\n\n")
            
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("".join(md_content))
            
        print(f"Saved to {output_path}")
        
    except Exception as e:
        print(f"Error converting {filename}: {e}")

def main():
    ensure_dir(OUTPUT_DIR)
    
    if not os.path.exists(SOURCE_DIR):
        print(f"Source directory not found: {SOURCE_DIR}")
        return

    files = [f for f in os.listdir(SOURCE_DIR) if f.endswith(".xlsx") and not f.startswith("~$")]
    
    if not files:
        print("No Excel files found.")
        return
        
    for f in files:
        convert_file(os.path.join(SOURCE_DIR, f))
        
if __name__ == "__main__":
    main()
